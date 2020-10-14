# coding=utf-8
import pygame
from pygame.draw import *
from random import randint
import openpyxl

RED = (255, 0, 0)
BLUE = (0, 0, 255)
YELLOW = (255, 255, 0)
GREEN = (0, 255, 0)
MAGENTA = (255, 0, 255)
CYAN = (0, 255, 255)
BLACK = (0, 0, 0)
COLORS = [RED, BLUE, YELLOW, GREEN, MAGENTA, CYAN]


def music():
    """
    включает выбранную музыку
    :param n: задаёт конкретный файл песни
    :return: none
    """
    pygame.init()
    pygame.mixer.init()
    if difficulty == 1:
        pygame.mixer.music.load('Music/Ievan Polkka.wav')
    if difficulty == 2:
        pygame.mixer.music.load('Music/Rollin.wav')
    if difficulty == 3:
        pygame.mixer.music.load('Music/Toxicity.wav')
    pygame.mixer.music.play(-1)


class Ball:

    def __init__(self, cor):
        self.x = cor[0]
        self.y = cor[1]
        self.r = cor[2]
        self.v_x = cor[3]
        self.v_y = cor[4]
        self.a_x = cor[5]
        self.a_y = cor[6]
        self.complexity = cor[7]
        self.color = self.colour()

    def draw_ball(self, x, y):
        """Redraws ball in x, y coordinates"""
        self.x = x
        self.y = y
        if abs(self.v_x) > 8 * self.complexity:
            self.a_x *= -1
            self.a_x //= 2
        if abs(self.v_y) > 8 * self.complexity:
            self.a_y *= -1
            self.a_y //= 2
        self.check_walls()
        self.v_x += self.a_x
        self.v_y += self.a_y
        circle(screen, self.color, (self.x, self.y), self.r)

    def draw_new_ball(self, x, y, v_x, v_y, a_x, a_y):
        """Draws new ball with given coordinates velocity and acceleration"""
        self.x = x
        self.y = y
        self.v_x = v_x
        self.v_y = v_y
        self.a_x = a_x
        self.a_y = a_y
        self.a_x = randint(-self.complexity, self.complexity)
        self.a_y = randint(-self.complexity, self.complexity)
        self.v_x += self.a_x
        self.v_y += self.a_y
        self.check_walls()
        circle(screen, self.color, (self.x, self.y), self.r)

    def check_walls(self):
        """Creates wall reflection"""
        if self.x > screen_height:
            self.v_x *= -1
            self.v_y = randint(-7, 7)
        if self.x < 0:
            self.v_x *= -1
            self.v_y = randint(-7, 7)
        if self.y > screen_width:
            self.v_x = randint(-7, 7)
            self.v_y *= -1
        if self.y < 0:
            self.v_x = randint(-7, 7)
            self.v_y *= -1

    def ball_coordinates(self):
        return self.x, self.y, self.r

    def movement_x(self):
        return self.v_x

    def movement_y(self):
        return self.v_y

    def complex(self):
        return self.complexity

    @staticmethod
    def colour():
        color = COLORS[randint(0, 5)]
        return color


def coordinates(compl):
    """Creates random coordinates for the new ball depending on the ball's complexity"""
    x = randint(100, screen_height - 100)
    y = randint(100, screen_width - 100)
    r = randint(30 / compl, 90 / compl)
    v_x = randint(-2 * compl, 2 * compl)
    v_y = randint(-2 * compl, 2 * compl)
    a_x = randint(-1 * compl, 1 * compl)
    a_y = randint(-1 * compl, 1 * compl)
    complex = compl
    return [x, y, r, v_x, v_y, a_x, a_y, complex]


def hit(position):
    """Checks if click hited the ball"""
    for i in balls:
        if i.ball_coordinates()[2] ** 2 > (position[0] - i.ball_coordinates()[0]) ** 2 + (
                position[1] - i.ball_coordinates()[1]) ** 2:
            return [i.complex(), i]
    return [0, 0]


def new_balls(compl):
    """Creates bew balls depending on the level difficulty"""
    if compl == 1:
        a = Ball(coordinates(1))
        balls = [a]
    if compl == 2:
        a = Ball(coordinates(2))
        b = Ball(coordinates(2))
        c = Ball(coordinates(3))
        balls = [a, b, c]
    if compl == 3:
        a = Ball(coordinates(3))
        b = Ball(coordinates(2))
        c = Ball(coordinates(3))
        d = Ball(coordinates(2))
        e = Ball(coordinates(3))
        balls = [a, b, c, d, e]
    return balls


def special(x, y):
    """Draws special aim that costs 10 points"""
    rect(screen, RED, (x - 20, y - 20, 40, 40))


def write_leaders():
    """Writes leaderboards"""
    wb = openpyxl.load_workbook(filename='leaderboard.xlsx')
    sheet = wb['test']
    names = []
    scores = []
    i = 1
    for j in range(10):
        names.append(str(sheet.cell(row=i, column=1).value))
        scores.append(int(sheet.cell(row=i, column=2).value))
        i += 1

    names.append(your_name)
    scores.append(success - fail)
    leaders = dict(zip(scores, names))
    lead_keys = list(leaders.keys())
    lead_keys = sorted(lead_keys, reverse=True)
    lead_items = []
    for j in lead_keys:
        lead_items.append((leaders[j]))
    lead_keys.pop()
    lead_items.pop()

    i = 1
    for j in lead_items:
        sheet.cell(row=i, column=1).value = j
        i += 1

    i = 1
    for j in lead_keys:
        sheet.cell(row=i, column=2).value = j
        i += 1

    wb.save('leaderboard.xlsx')


success = 0
fail = 0
time = 0
spec_x = -1000
spec_y = -1000

difficulty = int(input("Difficulty 1, 2 or 3: "))

pygame.init()

FPS = 45
game_length = 30

screen_height = 1200
screen_width = 700
screen = pygame.display.set_mode((screen_height, screen_width))

balls = new_balls(difficulty)

pygame.display.update()
clock = pygame.time.Clock()
finished = False

music()

intro = 0
while intro < 300:
    intro += 1
    clock.tick(FPS)

while time < game_length * FPS and not finished:
    clock.tick(FPS)
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            finished = True
        elif event.type == pygame.MOUSEBUTTONDOWN:
            if hit(event.pos)[0] > 0:
                success += hit(event.pos)[0]
                coor_new = coordinates(hit(event.pos)[1].complex())
                coor_new.pop(2)
                coor_new.pop()
                hit(event.pos)[1].draw_new_ball(*coor_new)
            elif abs(event.pos[0] - spec_x) < 20 and abs(event.pos[1] - spec_y) < 20:
                spec_x = -1000
                spec_y = -1000
                success += 10
            else:
                fail += 1
    if time % (FPS * (10 - difficulty)) == 0:
        screen.fill(BLACK)
        balls = new_balls(difficulty)
        for i in balls:
            coor_new = coordinates(i.complex())
            coor_new.pop(2)
            coor_new.pop()
            i.draw_new_ball(*coor_new)
    elif time % 3 == 0:
        screen.fill(BLACK)
        if randint(1, 55) == 50:
            spec_x = randint(50, screen_height - 50)
            spec_y = randint(50, screen_width - 50)
            special(spec_x, spec_y)
        for i in balls:
            i.draw_ball(i.ball_coordinates()[0] + i.movement_x(), i.ball_coordinates()[1] + i.movement_y())
    pygame.display.update()
    if time % 75 != 0:
        special(spec_x, spec_y)
    else:
        spec_x = -1000
        spec_y = -1000
    time += 1

pygame.quit()
your_name = raw_input("Your name: ")
print("Your score: " + str(success - fail))

write_leaders()
